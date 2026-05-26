"""
Motor de cálculo do Simulador ODP — Privalia
"""
import pandas as pd
import numpy as np
from datetime import date, timedelta
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import warnings
warnings.filterwarnings('ignore')

from params import (FERIADOS, get_pcs_palete, get_fat_in,
                    get_fat_out, get_setor)

# ── Dias úteis ─────────────────────────────────────────────────────────────────
def dias_uteis(base: date, n: int) -> date:
    step = 1 if n >= 0 else -1
    restante = abs(n)
    d = base
    while restante > 0:
        d += timedelta(days=step)
        if d.weekday() < 5 and d not in FERIADOS:
            restante -= 1
    return d

# ── Curvas de venda ────────────────────────────────────────────────────────────
def carregar_curvas(wb) -> dict:
    ws = wb['Curva Venda ODP']
    curvas = {}
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, values_only=True):
        if not row[1] or not row[2]:
            continue
        cat = str(row[1]).strip()
        try:
            wd = int(row[2])
        except:
            continue
        props = [v for v in row[3:] if isinstance(v, (int, float))]
        if props:
            s = sum(props) or 1
            curvas[(cat, wd)] = [p/s for p in props]
    return curvas

def get_curva(cat: str, wd: int, curvas: dict) -> list:
    if (cat, wd) in curvas:
        return curvas[(cat, wd)]
    for wdf in [wd] + list(range(max(1, wd-3), wd+4)):
        if ('Standard', wdf) in curvas:
            c = curvas[('Standard', wdf)]
            return (c + [0]*(wd-len(c)))[:wd] if len(c) < wd else c[:wd]
    return [1/wd] * wd

# ── Leituras de abas ───────────────────────────────────────────────────────────
def ler_calendario(wb) -> pd.DataFrame:
    ws = wb['Calendário atualizado']
    dados = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[4]: continue
        try: id_ = int(float(str(row[4])))
        except: continue
        modelo = str(row[41] or '').strip()
        if modelo.upper() != 'ODP': continue
        sd, ed = row[6], row[7]
        if not sd or not ed: continue
        if hasattr(sd, 'date'): sd = sd.date()
        if hasattr(ed, 'date'): ed = ed.date()
        try: wd = int(float(row[8])) if row[8] else 1
        except: wd = 1
        dados.append({
            'id':        id_,
            'campanha':  str(row[5] or '').strip(),
            'sd':        sd,  'ed': ed,  'wd': wd,
            'status_sf': str(row[10] or '').strip(),
            'cd':        str(row[11] or '').strip(),
            'cat':       str(row[12] or '').strip(),
            'subcat':    str(row[13] or '').strip(),
            'gerencia':  str(row[42] or '').strip(),
            'forecast':  float(row[16]) if row[16] else 0,
            'estoque':   float(row[18]) if row[18] else 0,
        })
    df = pd.DataFrame(dados)
    if df.empty: return df
    # ← CORREÇÃO: uma linha por campanha ODP (ignora duplicatas por CNPJ)
    return df.drop_duplicates(subset='id', keep='first').reset_index(drop=True)

def ler_vendas(wb) -> pd.Series:
    """Retorna total de peças vendidas por id de campanha."""
    ws = wb['Base Vendas Dia']
    dados = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[5]: continue
        try: id_ = int(float(str(row[5])))
        except: continue
        dia = row[8]
        if hasattr(dia, 'date'): dia = dia.date()
        pecas = float(row[13]) if row[13] else 0
        dados.append({'id': id_, 'dia': dia, 'pecas': pecas})
    return pd.DataFrame(dados)

def ler_liquida(wb) -> pd.DataFrame:
    ws = wb['Liquida']
    dados = []
    for row in ws.iter_rows(min_row=9, max_row=ws.max_row, values_only=True):
        if not row[7]: continue
        try: id_ = int(float(str(row[7])))
        except: continue
        dados.append({
            'id':      id_,
            'estoque': float(row[2]) if row[2] else 0,
            'venda':   float(row[3]) if row[3] else 0,
            'sellout': float(row[8]) if row[8] else 0,
        })
    df = pd.DataFrame(dados)
    if df.empty: return df
    return df.drop_duplicates(subset='id', keep='first').set_index('id')

# ── Motor principal ────────────────────────────────────────────────────────────
def calcular(arquivo, data_ref: date = None,
             ajustes: dict = None) -> tuple:
    """
    Retorna (df_campanhas, curvas, data_ref)
    ajustes = {id: {'sd', 'ed', 'estoque', 'forecast'}}
    """
    wb = load_workbook(arquivo, data_only=True)

    ws_odp = wb['ODP']
    qa25 = ws_odp.cell(25, column_index_from_string('QA')).value
    if data_ref is None:
        data_ref = qa25.date() if hasattr(qa25, 'date') else date.today()

    curvas = carregar_curvas(wb)
    cal    = ler_calendario(wb)
    if cal.empty:
        return pd.DataFrame(), curvas, data_ref

    df_vendas = ler_vendas(wb)
    liq       = ler_liquida(wb)

    # Agrega venda real até D-1 por campanha
    if not df_vendas.empty:
        vr_series = (df_vendas[df_vendas['dia'] < data_ref]
                     .groupby('id')['pecas'].sum())
    else:
        vr_series = pd.Series(dtype=float)

    registros = []
    for _, c in cal.iterrows():
        id_  = c['id']
        aj   = (ajustes or {}).get(id_, {})

        sd       = aj.get('sd',       c['sd'])
        ed       = aj.get('ed',       c['ed'])
        wd       = aj.get('wd',       c['wd'])
        estoque  = aj.get('estoque',  c['estoque'])
        forecast = aj.get('forecast', c['forecast'])

        cat       = c['cat']
        pcs_pal   = get_pcs_palete(cat)
        fat_in_v  = get_fat_in(cat)
        fat_out_v = get_fat_out(cat)
        setor_dep = get_setor(cat)
        pallets   = round(estoque / pcs_pal) if pcs_pal > 0 else 0

        dn  = dias_uteis(sd, -2)
        do_ = dias_uteis(ed,  3)

        # Status operacional
        if ed < data_ref:        status_op = 'x'
        elif sd <= data_ref <= ed: status_op = 'ON'
        else:                    status_op = '-'

        # Curva normalizada para wd dias
        curva = get_curva(cat, wd, curvas)
        if len(curva) < wd:  curva = curva + [0]*(wd-len(curva))
        elif len(curva) > wd: curva = curva[:wd]
        tot = sum(curva) or 1
        curva = [v/tot for v in curva]

        dias_decor = max(0, min((data_ref - sd).days, wd))
        dias_rest  = max(0, (ed - data_ref).days)
        dia_atual  = dias_decor + 1 if status_op == 'ON' else (wd if status_op == 'x' else 0)

        # Venda real
        vr = float(vr_series.get(id_, 0))
        if vr == 0 and id_ in liq.index:
            vr = float(liq.loc[id_, 'venda'])

        # SO congelado
        so_cong = float(liq.loc[id_, 'sellout']) if id_ in liq.index else 0
        if so_cong == 0 and estoque > 0:
            so_cong = forecast / estoque

        # QB→QR
        proj_acum    = sum(curva[:dias_decor])
        so_prev_acum = proj_acum
        so_real_acum = vr / forecast if forecast > 0 else 0
        desvio_abs   = so_real_acum - so_prev_acum
        desvio_rel   = (so_real_acum / so_prev_acum - 1) if so_prev_acum > 0 else 0
        desvio_med   = (desvio_abs + desvio_rel) / 2
        fator_ajuste = max(0.1, 1 + desvio_med)
        prop_rest    = 1 - proj_acum
        vol_adic     = min(prop_rest * fator_ajuste * forecast,
                           max(0, estoque - vr))
        reforecast   = vr + vol_adic
        mudanca_fc   = reforecast / forecast - 1 if forecast > 0 else 0
        novo_so      = reforecast / estoque if estoque > 0 else 0
        desvio_so    = novo_so - so_cong

        registros.append({
            'id': id_, 'campanha': c['campanha'],
            'status_op': status_op, 'status_sf': c['status_sf'],
            'sd': sd, 'ed': ed, 'wd': wd,
            'cat': cat, 'subcat': c['subcat'],
            'setor': setor_dep, 'cd': c['cd'], 'gerencia': c['gerencia'],
            'estoque': estoque, 'pallets': pallets,
            'pcs_pal': pcs_pal, 'fat_in': fat_in_v, 'fat_out': fat_out_v,
            'forecast': forecast, 'so_cong': so_cong,
            'dn': dn, 'do_': do_,
            'dias_decor': dias_decor, 'dias_rest': dias_rest, 'dia_atual': dia_atual,
            'proj_acum': proj_acum, 'so_prev_acum': so_prev_acum,
            'venda_real': vr, 'so_real_acum': so_real_acum,
            'desvio_abs': desvio_abs, 'desvio_rel': desvio_rel,
            'fator_ajuste': fator_ajuste, 'reforecast': reforecast,
            'novo_so': novo_so, 'mudanca_fc': mudanca_fc, 'desvio_so': desvio_so,
            'data_ref': data_ref, 'ajustada': bool(aj),
        })

    df = pd.DataFrame(registros)
    df['sd']  = pd.to_datetime(df['sd'])
    df['ed']  = pd.to_datetime(df['ed'])
    df['dn']  = pd.to_datetime(df['dn'])
    df['do_'] = pd.to_datetime(df['do_'])
    return df, curvas, data_ref

# ── Séries temporais ───────────────────────────────────────────────────────────
def serie_blocado(df, dt_ini, dt_fim, unidade='pallets'):
    rows = []
    for dt in pd.date_range(dt_ini, dt_fim, freq='D'):
        dt_d = dt.date()
        mask = (df['dn'].dt.date <= dt_d) & (df['do_'].dt.date >= dt_d)
        ativas = df[mask]
        if unidade == 'pallets':     vol = ativas['pallets'].sum()
        elif unidade == 'pecas_brutas': vol = ativas['estoque'].sum()
        else:                        vol = (ativas['estoque']/ativas['fat_in'].replace(0,1)).sum()
        rows.append({'data': dt, 'volume': round(vol),
                     'n_campanhas': len(ativas),
                     'campanhas': list(ativas['campanha'])})
    return pd.DataFrame(rows)

def serie_saidas(df, dt_ini, dt_fim, curvas, unidade='pecas_conv'):
    rows = []
    for _, camp in df.iterrows():
        sd  = camp['sd'].date()
        ed  = camp['ed'].date()
        wd  = int(camp['wd'])
        fat = camp['fat_out'] if unidade == 'pecas_conv' else 1.0
        data_ref = camp['data_ref']

        curva = get_curva(camp['cat'], wd, curvas)
        if len(curva) < wd:  curva = curva + [0]*(wd-len(curva))
        elif len(curva) > wd: curva = curva[:wd]
        tot = sum(curva) or 1
        curva = [v/tot for v in curva]

        decor = max(0, min((data_ref - sd).days, wd))
        tot_prev_real = sum(curva[:decor]) or 1

        for i in range(wd):
            dt_d = sd + timedelta(days=i)
            if dt_d < dt_ini or dt_d > dt_fim: continue
            prop = curva[i]
            if dt_d < data_ref:
                vol_b = (prop / tot_prev_real) * camp['venda_real'] if decor > 0 else 0
                tipo  = 'realizado'
            else:
                vol_b = prop * camp['reforecast']
                tipo  = 'reforecast' if camp['status_op'] == 'ON' else 'previsto'
            vol = vol_b / fat if fat > 0 else vol_b
            rows.append({
                'data': pd.Timestamp(dt_d), 'id': camp['id'],
                'campanha': camp['campanha'], 'cat': camp['cat'],
                'setor': camp['setor'], 'cd': camp['cd'],
                'status': camp['status_op'], 'volume': max(0, round(vol)),
                'tipo': tipo,
            })
    return pd.DataFrame(rows)

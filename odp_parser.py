"""
Módulo de parsing do Simulador ODP — Privalia
Lê diretamente o .xlsm e reconstrói os dados estruturados
sem depender de fórmulas calculadas pelo Excel.
"""
import pandas as pd
import numpy as np
from datetime import date, timedelta
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import warnings
warnings.filterwarnings('ignore')

# ─── Mapeamento de colunas da aba ODP ────────────────────────────────────────
COLS_ODP = {
    'id':        'A',  'campanha':    'B',  'status':   'C',
    'sd':        'D',  'ed':          'E',  'wd':       'F',
    'class_':    'G',  'cd':          'H',  'cat':      'I',
    'setor':     'J',  'dep_setor':   'K',
    'pcs_pal':   'M',  'fat_in':      'N',  'fat_out':  'O',
    'dn':        'DN', 'do_':         'DO',
    'estoque':   'DQ', 'pallets':     'DR', 'ttm':      'DS',
    'forecast':  'IF', 'so_cong':     'IG',
}

# ─── Feriados e dias úteis ────────────────────────────────────────────────────
FERIADOS = {
    date(2026, 1, 1), date(2026, 4, 3), date(2026, 4, 21),
    date(2026, 5, 1), date(2026, 6, 11), date(2026, 9, 7),
    date(2026, 10, 12), date(2026, 11, 2), date(2026, 11, 15),
    date(2026, 12, 25),
}

def dias_uteis(data_base: date, n: int, feriados=FERIADOS) -> date:
    """Avança/recua n dias úteis (positivo=futuro, negativo=passado)."""
    step = 1 if n >= 0 else -1
    restante = abs(n)
    d = data_base
    while restante > 0:
        d += timedelta(days=step)
        if d.weekday() < 5 and d not in feriados:
            restante -= 1
    return d

# ─── Curvas de venda ODP ──────────────────────────────────────────────────────
def carregar_curvas(wb) -> dict:
    """Retorna dict {(categoria, webdays): [d1, d2, ...d_n]} com proporções diárias."""
    ws = wb['Curva Venda ODP']
    curvas = {}
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, values_only=True):
        if not row[1] or not row[2]:
            continue
        cat = str(row[1]).strip()
        try:
            wd = int(row[2])
        except:
            continue
        props = [v for v in row[3:] if isinstance(v, (int, float))]
        if props:
            curvas[(cat, wd)] = props
    return curvas

def get_curva(cat: str, wd: int, curvas: dict) -> list:
    """Retorna curva para categoria+webdays, com fallback para Standard."""
    if (cat, wd) in curvas:
        return curvas[(cat, wd)]
    # fallback standard
    for wdf in range(wd, 0, -1):
        if ('Standard', wdf) in curvas:
            return curvas[('Standard', wdf)]
    # último recurso: distribuição uniforme
    return [1/wd] * wd

# ─── Parâmetros da aba Suporte ────────────────────────────────────────────────
def carregar_suporte(wb) -> dict:
    """Lê pcs/palete e fatores IN/OUT da aba Suporte."""
    ws = wb['Suporte']
    pcs_pal = {}
    fat_in = {}
    fat_out = {}
    for row in ws.iter_rows(min_row=2, max_row=200, values_only=True):
        if row[0] and row[2]:  # col A=categoria, C=pcs/palete
            try:
                pcs_pal[str(row[0]).strip()] = float(row[2])
            except:
                pass
        if row[4] and row[5] and row[6]:  # col E=cat, F=fat_in, G=fat_out
            try:
                fat_in[str(row[4]).strip()] = float(row[5])
                fat_out[str(row[4]).strip()] = float(row[6])
            except:
                pass
    return {'pcs_pal': pcs_pal, 'fat_in': fat_in, 'fat_out': fat_out}

# ─── Leitura principal da aba ODP ─────────────────────────────────────────────
def carregar_odp(arquivo, data_ref: date = None) -> pd.DataFrame:
    """
    Lê o .xlsm e retorna DataFrame com todas as campanhas ODP reais.
    Recalcula: DN, DO, pallets, reforecast (QB→QR).
    """
    wb = load_workbook(arquivo, data_only=True)
    ws = wb['ODP']

    # índices numéricos das colunas
    col_idx = {nome: column_index_from_string(col) for nome, col in COLS_ODP.items()}

    # data de referência = QA25 ou hoje
    qa25 = ws.cell(row=25, column=column_index_from_string('QA')).value
    if data_ref is None:
        data_ref = qa25.date() if hasattr(qa25, 'date') else date.today()

    curvas = carregar_curvas(wb)
    sup    = carregar_suporte(wb)

    registros = []
    for row_num in range(27, ws.max_row + 1):
        id_val = ws.cell(row=row_num, column=1).value
        if id_val is None:
            continue
        try:
            id_num = int(float(str(id_val)))
        except:
            continue

        # filtra IDs de campanhas reais (>= 200000, exclui placeholders 111111 etc)
        if id_num < 200000:
            continue

        def v(nome):
            val = ws.cell(row=row_num, column=col_idx[nome]).value
            return val

        status   = str(v('status') or '').strip()
        campanha = str(v('campanha') or '').strip()
        cat      = str(v('cat') or '').strip()
        cd       = str(v('cd') or '').strip()
        setor    = str(v('setor') or '').strip()
        class_   = str(v('class_') or '').strip()

        # datas
        def to_date(val):
            if val is None: return None
            if hasattr(val, 'date'): return val.date()
            if isinstance(val, date): return val
            return None

        sd       = to_date(v('sd'))
        ed       = to_date(v('ed'))
        dn_raw   = to_date(v('dn'))
        do_raw   = to_date(v('do_'))

        if sd is None or ed is None:
            continue

        # webdays e numéricos
        def fnum(nome, default=0):
            val = v(nome)
            try: return float(val) if val not in (None, 'n/a', '#DIV/0!', 'x', '-', '') else default
            except: return default

        wd       = int(fnum('wd', 1))
        estoque  = fnum('estoque')
        pcs_pal  = fnum('pcs_pal') or sup['pcs_pal'].get(cat, 150)
        fat_in   = fnum('fat_in')  or sup['fat_in'].get(cat, 1.0)
        fat_out  = fnum('fat_out') or sup['fat_out'].get(cat, 1.0)
        forecast = fnum('forecast')
        so_cong  = fnum('so_cong')

        # recalcula pallets se inválido
        pallets = fnum('pallets')
        if pallets <= 0 and pcs_pal > 0:
            pallets = round(estoque / pcs_pal)

        # datas de blocado — usa o lido se válido, senão recalcula
        if dn_raw and dn_raw > date(2020, 1, 1):
            dn = dn_raw
        else:
            dn = dias_uteis(sd, -2)
        if do_raw and do_raw > date(2020, 1, 1):
            do_ = do_raw
        else:
            do_ = dias_uteis(ed, 3)

        # ── Racional de reforecast (QB→QR) ──────────────────────────────────
        curva = get_curva(cat, wd, curvas)
        dias_venda    = (data_ref - sd).days if sd <= data_ref else 0
        dias_rest     = max(0, (ed - data_ref).days)
        dias_decor    = min(dias_venda, wd)  # dias dentro do período de venda

        # projeção acumulada até D-1 (soma da curva até o dia atual)
        proj_acum = sum(curva[:dias_decor]) if curva else 0
        so_prev_acum = proj_acum  # como % do estoque

        # venda real acumulada — lida diretamente (QF)
        venda_real_raw = ws.cell(row=row_num, column=column_index_from_string('QF')).value
        try:
            venda_real = float(venda_real_raw) if venda_real_raw not in (None, 'n/a', 'x', '-', '') else 0
        except:
            venda_real = 0

        so_real_acum = venda_real / forecast if forecast > 0 else 0

        # desvio e fator de ajuste
        desvio_abs = so_real_acum - so_prev_acum
        desvio_rel = (so_real_acum / so_prev_acum - 1) if so_prev_acum > 0 else 0
        desvio_medio = (desvio_abs + desvio_rel) / 2
        fator_ajuste = 1 + desvio_medio

        # projeção do restante
        prop_restante = 1 - so_prev_acum
        proj_adic = prop_restante * fator_ajuste
        volume_adic = min(proj_adic * forecast, estoque - venda_real)
        volume_adic = max(0, volume_adic)

        reforecast = venda_real + volume_adic
        novo_so = reforecast / estoque if estoque > 0 else 0
        mudanca_vs_fc = reforecast / forecast - 1 if forecast > 0 else 0
        desvio_vs_so = novo_so - so_cong

        # dia atual da campanha (D1=1)
        dia_atual = max(0, min(dias_decor + 1, wd))

        registros.append({
            'id':           id_num,
            'campanha':     campanha,
            'status':       status,
            'sd':           sd,
            'ed':           ed,
            'wd':           wd,
            'cat':          cat,
            'setor':        setor,
            'class_':       class_,
            'cd':           cd,
            'estoque':      estoque,
            'pallets':      pallets,
            'pcs_pal':      pcs_pal,
            'fat_in':       fat_in,
            'fat_out':      fat_out,
            'forecast':     forecast,
            'so_cong':      so_cong,
            'dn':           dn,
            'do_':          do_,
            # reforecast
            'dias_venda':   dias_venda,
            'dias_rest':    dias_rest,
            'dia_atual':    dia_atual,
            'proj_acum':    proj_acum,
            'so_prev_acum': so_prev_acum,
            'venda_real':   venda_real,
            'so_real_acum': so_real_acum,
            'desvio_abs':   desvio_abs,
            'desvio_rel':   desvio_rel,
            'fator_ajuste': fator_ajuste,
            'reforecast':   reforecast,
            'novo_so':      novo_so,
            'mudanca_vs_fc':mudanca_vs_fc,
            'desvio_vs_so': desvio_vs_so,
            'data_ref':     data_ref,
        })

    df = pd.DataFrame(registros)
    if df.empty:
        return df

    df['sd']  = pd.to_datetime(df['sd'])
    df['ed']  = pd.to_datetime(df['ed'])
    df['dn']  = pd.to_datetime(df['dn'])
    df['do_'] = pd.to_datetime(df['do_'])

    return df


# ─── Gerador de série temporal de blocado ─────────────────────────────────────
def gerar_serie_blocado(df: pd.DataFrame, dt_ini: date, dt_fim: date) -> pd.DataFrame:
    """
    Para cada dia no intervalo, soma os pallets de campanhas
    cujo blocado (dn → do_) cobre aquele dia.
    Retorna DataFrame: data | pallets_total | [por campanha]
    """
    datas = pd.date_range(dt_ini, dt_fim, freq='D')
    rows = []
    for dt in datas:
        dt_d = dt.date()
        ativas = df[(df['dn'].dt.date <= dt_d) & (df['do_'].dt.date >= dt_d)]
        rows.append({
            'data': dt,
            'pallets': ativas['pallets'].sum(),
            'n_campanhas': len(ativas),
        })
    return pd.DataFrame(rows)


# ─── Gerador de série temporal de saídas ─────────────────────────────────────
def gerar_serie_saidas(df: pd.DataFrame, dt_ini: date, dt_fim: date,
                        unidade: str = 'convertido') -> pd.DataFrame:
    """
    Para cada campanha e cada dia do intervalo, distribui o volume
    de saída pela curva de venda (reforecast ou previsão).
    unidade: 'bruto' | 'convertido'
    Retorna DataFrame: data | campanha | volume | tipo
    """
    from openpyxl import load_workbook  # curvas já carregadas via df
    rows = []
    datas = pd.date_range(dt_ini, dt_fim, freq='D')

    for _, camp in df.iterrows():
        sd = camp['sd'].date()
        ed = camp['ed'].date()
        wd = int(camp['wd'])
        fat = camp['fat_out'] if unidade == 'convertido' else 1.0
        data_ref = camp['data_ref']

        # reconstrói a curva básica (sem recarregar wb — usa proporção uniforme como fallback)
        # O app.py real vai passar as curvas carregadas; aqui simplificamos
        if wd > 0:
            curva = [1/wd] * wd
        else:
            continue

        for i, dt in enumerate(datas):
            dt_d = dt.date()
            if dt_d < sd or dt_d > ed:
                continue

            dia_rel = (dt_d - sd).days  # 0-indexed
            if dia_rel >= len(curva):
                continue

            prop = curva[dia_rel]
            vol_bruto = prop * camp['reforecast']
            vol = vol_bruto / fat if fat > 0 else vol_bruto

            tipo = 'realizado' if dt_d < data_ref else (
                   'reforecast' if camp['status'] == 'ON' else 'previsto')

            rows.append({
                'data':     dt,
                'campanha': camp['campanha'],
                'id':       camp['id'],
                'cat':      camp['cat'],
                'cd':       camp['cd'],
                'status':   camp['status'],
                'volume':   round(vol),
                'tipo':     tipo,
            })

    return pd.DataFrame(rows)
